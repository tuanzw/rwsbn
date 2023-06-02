import os
from pathlib import PurePath
from glob import glob

from openpyxl import load_workbook

import shortuuid
import uuid

from logapi import logger


class ExcelDispatcher:
    def __init__(
        self,
        s_split,
        attachment_folder,
        attchement_move_to_folder,
        worksheet_name,
        trip_col_idx,
        tn_col_idx,
        bn_col_idx,
        first_row_idx,
    ):
        self._s_split = s_split
        self._attachment_folder = attachment_folder
        self._attchement_move_to_folder = attchement_move_to_folder
        self._worksheet_name = worksheet_name
        self._trip_col_idx = trip_col_idx
        self._tn_col_idx = tn_col_idx
        self._bn_col_idx = bn_col_idx
        self._first_row_idx = first_row_idx

    def booking_number(self):
        u = uuid.uuid4()
        s = shortuuid.encode(u)
        short = s[:8]
        return short

    def xstr(self, s):
        return "" if s is None else str(s).strip().upper()

    def proceed_excel(self):
        # list of .xlsx file paths excluding [~$] temporiraly opening xlxs file
        file_paths = glob(f"{self._attachment_folder}[!~$]*.xlsx")
        for file_path in file_paths:
            try:
                self.fill_booking_number(file_path)
                logger.info(file_path)
            except PermissionError as p:
                # [Errno 13] Permission denied: -> file is locked for opening
                logger.debug(p)
            except KeyError as k:
                # Worksheet Template does not exist. -> no "Template" worksheet in file
                logger.debug(f"{k} for file {file_path}")

    def get_existing_booking_number(self, file_name) -> dict:
        archive_path = PurePath(
            self._attachment_folder, self._attchement_move_to_folder, file_name
        )
        existing_bn_dict = {}
        if os.path.isfile(archive_path):
            wb = load_workbook(filename=archive_path)
            ws = wb[self._worksheet_name]

            for data in ws.iter_rows(
                min_row=self._first_row_idx,
                max_row=ws.max_row,
                min_col=self._trip_col_idx,
                max_col=self._bn_col_idx,
            ):
                # no truck number or no booking number filled, skip
                if not self.xstr(
                    data[self._tn_col_idx - self._trip_col_idx].value
                ) or not self.xstr(data[self._bn_col_idx - self._trip_col_idx].value):
                    continue
                # key of dict is Trip-TruckNumber, example: 1-60LD04204
                key = f"{self.xstr(data[0].value)}{self._s_split}{self.xstr(data[self._tn_col_idx-self._trip_col_idx].value)}"
                existing_bn_dict[key] = data[
                    self._bn_col_idx - self._trip_col_idx
                ].value
        logger.info(f"__EXISTING BKNM__{existing_bn_dict}")
        return existing_bn_dict

    def fill_booking_number(self, file_path):
        wb = load_workbook(filename=file_path)

        ws = wb[self._worksheet_name]
        # print(ws.calculate_dimension())

        bn_dict = self.get_existing_booking_number(os.path.basename(file_path))
        for data in ws.iter_rows(
            min_row=self._first_row_idx,
            max_row=ws.max_row,
            min_col=self._trip_col_idx,
            max_col=self._bn_col_idx,
        ):
            # no truck number filled, skip
            if not self.xstr(data[self._tn_col_idx - self._trip_col_idx].value):
                continue
            # key of dict is Trip-TruckNumber, example: 1-60LD04204
            key = f"{self.xstr(data[0].value)}{self._s_split}{self.xstr(data[self._tn_col_idx-self._trip_col_idx].value)}"
            # generate booking number for the trip/truck
            if bn_dict.get(key) is None:
                bn_dict[key] = self.booking_number()
            # fill booking number into column Booking Number in excel file
            data[self._bn_col_idx - self._trip_col_idx].value = bn_dict.get(key)

        logger.info(f"__FILLING  BKNM__{bn_dict}")
        wb.save(filename=file_path)

    def get_bn_dict_from_file(self, file_path):
        wb = load_workbook(filename=file_path, read_only=True)

        ws = wb[self._worksheet_name]
        bn_dict = {}
        for data in ws.iter_rows(
            min_row=self._first_row_idx,
            max_row=ws.max_row,
            min_col=self._trip_col_idx,
            max_col=self._bn_col_idx,
        ):
            # no truck number filled, skip
            if not self.xstr(data[self._tn_col_idx - self._trip_col_idx].value):
                continue
            # key of dict is Trip-TruckNumber, example: 1-60LD04204
            key = f"{self.xstr(data[0].value)}{self._s_split}{self.xstr(data[self._tn_col_idx-self._trip_col_idx].value)}"
            bn_dict[key] = data[self._bn_col_idx - self._trip_col_idx].value

        wb.close()
        return bn_dict
